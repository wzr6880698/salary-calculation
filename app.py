import openpyxl
from openpyxl import load_workbook
import os
from datetime import datetime, timedelta
import re
import streamlit as st
from io import BytesIO
import tempfile

# ============================
# 原有的核心逻辑类 (保持不变)
# ============================

class DateParser:
    @staticmethod
    def parse(date_str):
        if not date_str:
            return None
        if isinstance(date_str, (int, float)):
            return DateParser._parse_excel_number(date_str)
        date_str = str(date_str).strip()
        formats = [
            (r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", "%Y-%m-%d"),
            (r"\d{4}年\d{1,2}月\d{1,2}日", "%Y年%m月%d日"),
            (r"\d{2}年\d{1,2}月\d{1,2}日", "%y年%m月%d日"),
            (r"\d{1,2}月\d{1,2}日", "%m月%d日"),
            (r"\d{4}-\d{1,2}-\d{1,2} \d{1,2}:\d{1,2}:\d{1,2}", "%Y-%m-%d")
        ]
        for pattern, fmt in formats:
            if re.match(pattern, date_str):
                try:
                    dt = datetime.strptime(date_str.split()[0] if " " in date_str else date_str, fmt)
                    return dt.strftime("%Y/%m/%d")
                except ValueError:
                    continue
        return None

    @staticmethod
    def _parse_excel_number(num):
        try:
            base_date = datetime(1899, 12, 30)
            delta = timedelta(days=int(num))
            return (base_date + delta).strftime("%Y/%m/%d")
        except (ValueError, TypeError):
            return None


class DataValidator:
    @staticmethod
    def is_valid_name(name):
        if not name or not isinstance(name, str):
            return False
        name = name.strip()
        return (name and len(name) >= 2 and
                name not in ["姓名", "合计", "序号", None, "日期", "优萌宠物车间生产日报表", "生产日报表"])

    @staticmethod
    def is_valid_number(value):
        try:
            float(value)
            return True
        except (TypeError, ValueError):
            return False

    @staticmethod
    def validate_record(record):
        required_fields = ["日期", "姓名", "产品名称"]
        for field in required_fields:
            if not record.get(field):
                return False
        return True


class WorkshopDataExtractor:
    def __init__(self, sheet_name):
        self.sheet_name = sheet_name
        self.current_date = None
        self.current_batch = "0"
        self.current_products = []
        self.headers = []
        self.date_pattern = re.compile(r'(\d{4}[-/年]\d{1,2}[-/月]\d{1,2}日?)')

    def extract(self, ws, data_list):
        print(f"\n[{self.sheet_name}车间] 开始处理工作表")
        self._find_initial_metadata(ws)
        for row in ws.iter_rows():
            self._process_row(row, data_list)

    def _find_initial_metadata(self, ws):
        for row in ws.iter_rows(max_row=10):
            for cell in row:
                if cell.value:
                    date_match = self.date_pattern.search(str(cell.value))
                    if date_match:
                        self.current_date = DateParser.parse(date_match.group())
                    if isinstance(cell.value, str):
                        cell_value = str(cell.value).strip()
                        if '批次号：' in cell_value:
                            self.current_batch = cell_value.split('：', 1)[-1].strip() or "0"
                        elif '批号：' in cell_value:
                            self.current_batch = cell_value.split('：', 1)[-1].strip() or "0"

    def _try_extract_metadata_from_row(self, row):
        for cell in row:
            if cell.value:
                parsed_date = DateParser.parse(cell.value)
                if parsed_date:
                    self.current_date = parsed_date
                if isinstance(cell.value, str):
                    cell_value = str(cell.value).strip()
                    if '批次号：' in cell_value:
                        self.current_batch = cell_value.split('：', 1)[-1].strip() or "0"
                    elif '批号：' in cell_value:
                        self.current_batch = cell_value.split('：', 1)[-1].strip() or "0"

    def _process_row(self, row, data_list):
        raise NotImplementedError

    def _create_record(self, name, product, quantity, price, amount, batch=None, note=""):
        record = {
            "日期": self.current_date,
            "姓名": name,
            "批次号": batch if batch is not None else self.current_batch,
            "产品名称": product,
            "数量": float(quantity) if quantity is not None and DataValidator.is_valid_number(quantity) else 0,
            "计量单位": "",
            "单价": float(price) if price is not None and DataValidator.is_valid_number(price) else 0,
            "金额": float(amount) if amount is not None and DataValidator.is_valid_number(amount) else 0,
            "车间名称": self.sheet_name,
            "备注": note
        }
        if record["金额"] == 0 and record["数量"] and record["单价"]:
            record["金额"] = record["数量"] * record["单价"]
        return record if DataValidator.validate_record(record) else None


class RaorouExtractor(WorkshopDataExtractor):
    def _is_header_row(self, row):
        if len(row) > 1 and DataValidator.is_valid_name(row[1].value):
            return False
        for cell in row:
            if cell.value:
                cell_value = str(cell.value).strip()
                if any(keyword in cell_value for keyword in ["数量", "单价", "金额", "件数", "价格", "总价", "备注"]):
                    return True
        return False

    def _parse_header_row(self, row):
        print(f"  [调试-Raorou] 发现表头行，进行解析。")
        self.current_products = []
        self.headers = []

        quantity_cols = []
        for i, cell in enumerate(row):
            if cell.value:
                cell_value = str(cell.value).strip()
                if any(keyword in cell_value for keyword in ["数量", "件数"]):
                    quantity_cols.append(i)

        for q_col in quantity_cols:
            batch_cell = row[q_col].offset(row=-1) if row[q_col].row > 1 else None
            batch = batch_cell.value if batch_cell else "0"
            batch = str(batch).strip() if batch else "0"

            product_cell = row[q_col].offset(row=-1, column=1) if row[q_col].row > 1 and q_col + 1 < len(row) else None
            product = product_cell.value if product_cell else None
            product = str(product).strip() if product else f"产品{len(self.current_products) + 1}"

            if product not in self.current_products:
                self.current_products.append(product)

            price_col = None
            amount_col = None
            note_col = None

            for j in range(q_col + 1, min(q_col + 5, len(row))):
                if j < len(row) and row[j].value:
                    cell_value = str(row[j].value).strip()
                    if any(keyword in cell_value for keyword in ["单价", "价格"]):
                        price_col = j
                    elif any(keyword in cell_value for keyword in ["金额", "总价"]):
                        amount_col = j
                    elif "备注" in cell_value:
                        note_col = j

            if price_col is None:
                price_col = q_col + 1 if q_col + 1 < len(row) else None

            if amount_col is None:
                amount_col = q_col + 2 if q_col + 2 < len(row) else None

            if note_col is None:
                note_col = q_col + 3 if q_col + 3 < len(row) else None

            self.headers.append({
                'col': q_col + 1,
                'type': '数量',
                'product': product,
                'batch': batch
            })

            if price_col is not None:
                self.headers.append({
                    'col': price_col + 1,
                    'type': '单价',
                    'product': product
                })

            if amount_col is not None:
                self.headers.append({
                    'col': amount_col + 1,
                    'type': '金额',
                    'product': product
                })

            if note_col is not None:
                self.headers.append({
                    'col': note_col + 1,
                    'type': '备注',
                    'product': product
                })

    @staticmethod
    def _find_product_names(sheet, current_row):
        products = []
        for r in range(max(1, current_row - 5), current_row):
            for cell in sheet[r]:
                if cell.value and isinstance(cell.value, str):
                    cell_value = str(cell.value).strip()
                    if '产品名称：' in cell_value:
                        product = cell_value.split('：', 1)[-1].strip()
                        if product and product not in products:
                            products.append(product)
                    elif '品名：' in cell_value:
                        product = cell_value.split('：', 1)[-1].strip()
                        if product and product not in products:
                            products.append(product)

        if len(products) < 3:
            products = ["5\"*12g漂白皮卷绕鸭肉", "螺旋三明治", "拆钩子"]

        return products

    def _is_data_row(self, row):
        return len(row) > 1 and DataValidator.is_valid_name(row[1].value)

    def _parse_data_row(self, row, data_list):
        name = row[1].value if len(row) > 1 else None
        if not name or not DataValidator.is_valid_name(name):
            return

        for i in range(0, len(self.headers), 4):
            if i + 3 >= len(self.headers):
                continue

            product_info = self.headers[i]
            product = product_info['product']
            batch = product_info.get('batch', "0")

            qty_col = self.headers[i]['col'] - 1
            price_col = self.headers[i + 1]['col'] - 1 if i + 1 < len(self.headers) else None
            amount_col = self.headers[i + 2]['col'] - 1 if i + 2 < len(self.headers) else None
            note_col = self.headers[i + 3]['col'] - 1 if i + 3 < len(self.headers) else None

            qty = row[qty_col].value if qty_col < len(row) else None
            price = row[price_col].value if price_col is not None and price_col < len(row) else None
            amount = row[amount_col].value if amount_col is not None and amount_col < len(row) else None

            note = ""
            if note_col is not None and note_col < len(row):
                try:
                    note_cell = row[note_col]
                    note = note_cell.value
                except Exception:
                    note = ""

            has_data = False

            if qty is not None:
                qty_str = str(qty).strip()
                if qty_str != "" and DataValidator.is_valid_number(qty_str):
                    has_data = True

            if not has_data and amount is not None:
                amount_str = str(amount).strip()
                if amount_str != "" and DataValidator.is_valid_number(amount_str):
                    has_data = True

            if not has_data and note is not None:
                try:
                    note_str = str(note).strip()
                    if note_str != "":
                        has_data = True
                except Exception as e:
                    print(f"  [警告] 处理备注 '{note}' 时出错: {str(e)}，将忽略此备注信息。")

            if has_data:
                record = self._create_record(
                    name,
                    product,
                    qty if qty is not None else 0,
                    price if price is not None else 0,
                    amount if amount is not None else 0,
                    batch,
                    str(note) if note is not None else ""
                )
                if record:
                    data_list.append(record)
                    # print(f"  提取记录: {record}") # 注释掉以避免网页刷屏

    def _process_row(self, row, data_list):
        if self._is_header_row(row):
            self._parse_header_row(row)
        elif self._is_data_row(row):
            # name = row[1].value if len(row) > 1 else "未知"
            # print(f"  [调试-Raorou] 发现数据行，姓名: '{name}'，开始解析。")
            self._parse_data_row(row, data_list)
        else:
            self._try_extract_metadata_from_row(row)


class ZhizuoExtractor(RaorouExtractor):
    pass


class BaozhuangExtractor(WorkshopDataExtractor):
    def extract(self, ws, data_list):
        print(f"\n[{self.sheet_name}车间] 开始处理工作表")
        max_col = ws.max_column or len(list(ws.iter_rows())[0]) if ws.iter_rows() else 0

        for row in ws.iter_rows():
            if not any(cell.value for cell in row):
                continue
            self._process_row(row, data_list, max_col)

    def _is_header_row(self, row):
        return False

    def _parse_header_row(self, row):
        pass

    def _is_data_row(self, row):
        return False

    def _parse_data_row(self, row, data_list, max_col):
        block_size = 8
        block_count = (max_col + block_size - 1) // block_size

        for block_index in range(block_count):
            offset = block_index * block_size

            if offset >= len(row):
                continue

            name_col = offset + 1
            product_col = offset + 3

            if name_col >= len(row) or product_col >= len(row):
                continue

            name_cell = row[name_col]
            if not (name_cell.value and DataValidator.is_valid_name(name_cell.value)):
                continue

            product_cell = row[product_col]
            if not (product_cell.value and isinstance(product_cell.value, str) and
                    not any(keyword in str(product_cell.value) for keyword in ["产品名称", "品名"])):
                continue

            name = str(name_cell.value).strip()

            date_col = offset
            batch_col = offset + 2
            quantity_col = offset + 4
            price_col = offset + 5
            amount_col = offset + 6
            note_col = offset + 7

            if date_col < len(row) and row[date_col].value:
                parsed_date = DateParser.parse(row[date_col].value)
                if parsed_date:
                    self.current_date = parsed_date

            batch = row[batch_col].value if batch_col < len(row) else "0"
            product = row[product_col].value if product_col < len(row) else ""
            quantity = row[quantity_col].value if quantity_col < len(row) else 0
            price = row[price_col].value if price_col < len(row) else 0
            amount = row[amount_col].value if amount_col < len(row) else 0

            note = ""
            if note_col < len(row):
                try:
                    note_cell = row[note_col]
                    note = note_cell.value
                except Exception:
                    note = ""

            has_data = False
            if quantity is not None:
                try:
                    quantity_str = str(quantity).strip()
                    if quantity_str != "" and DataValidator.is_valid_number(quantity_str):
                        has_data = True
                except Exception:
                    pass

            if not has_data and amount is not None:
                try:
                    amount_str = str(amount).strip()
                    if amount_str != "" and DataValidator.is_valid_number(amount_str):
                        has_data = True
                except Exception:
                    pass

            if not has_data and note is not None:
                try:
                    note_str = str(note).strip()
                    if note_str != "":
                        has_data = True
                except Exception as e:
                    print(f"  [警告] 处理备注 '{note}' 时出错: {str(e)}，将忽略此备注信息。")

            if product and has_data:
                record = self._create_record(name, product, quantity, price, amount, batch,
                                             str(note) if note is not None else "")
                if record:
                    data_list.append(record)
                    # print(f"  提取记录: {record}")

    def _process_row(self, row, data_list, max_col=None):
        if max_col is None:
            max_col = len(row)
        self._try_extract_metadata_from_row(row)
        self._parse_data_row(row, data_list, max_col)


# ============================
# 适配 Streamlit 的输出函数
# ============================

def save_to_output(data_list):
    """将数据保存到内存中的 BytesIO 对象，而不是磁盘路径"""
    if not data_list:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "数据收集表"

    headers = ["日期", "姓名", "批次号", "产品名称", "数量", "计量单位", "单价", "金额", "车间名称", "备注"]

    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for i, data in enumerate(data_list, 2):
        for col, key in enumerate(headers, 1):
            value = data.get(key, "")
            ws.cell(row=i, column=col, value=value)

    # 将工作簿写入内存缓冲区
    output_buffer = BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0) # 将指针移回开头
    return output_buffer


# ============================
# Streamlit 界面与主逻辑
# ============================

def main():
    st.set_page_config(page_title="车间日报提取工具", layout="wide")
    st.title("🏭 车间生产日报数据处理系统")
    st.markdown("""
    **使用说明：**
    1. 点击下方按钮上传车间日报表文件（支持 .xlsx, .xls）。
    2. 系统会自动识别文件名中包含 "优萌车间" 或 "生产日报" 的文件。
    3. 点击 "开始处理" 按钮。
    4. 处理完成后，点击 "下载结果文件" 按钮保存汇总表。
    """)
    st.markdown("---")

    # 1. 文件上传组件 (替代原有的 input_folder_path)
    uploaded_files = st.file_uploader(
        "📤 请选择要处理的文件 (可多选)", 
        type=['xlsx', 'xls'], 
        accept_multiple_files=True
    )

    if st.button("🚀 开始处理", type="primary"):
        if not uploaded_files:
            st.warning("⚠️ 请先上传至少一个文件！")
        else:
            all_data = []
            progress_bar = st.progress(0)
            status_text = st.empty()
            processed_count = 0

            # 2. 处理逻辑 (替代原有的 os.listdir 遍历)
            for uploaded_file in uploaded_files:
                processed_count += 1
                progress_bar.progress(processed_count / len(uploaded_files))
                status_text.text(f"正在处理: {uploaded_file.name} ...")
                
                # 文件名过滤 (保持原有逻辑)
                if "优萌车间" not in uploaded_file.name and "生产日报" not in uploaded_file.name:
                    st.info(f"⏭️ 文件 '{uploaded_file.name}' 不包含关键字，已跳过。")
                    continue
                
                try:
                    # openpyxl 可以直接读取 UploadedFile 对象
                    # 但为了兼容性，我们使用临时文件方式读取
                    with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(uploaded_file.name)[1]) as tmp:
                        tmp.write(uploaded_file.getbuffer())
                        tmp_path = tmp.name
                    
                    wb = load_workbook(tmp_path, data_only=True)
                    
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        extractor = None
                        if "绕肉" in sheet_name:
                            extractor = RaorouExtractor(sheet_name)
                        elif "制作" in sheet_name:
                            extractor = ZhizuoExtractor(sheet_name)
                        elif "包装" in sheet_name or "挑选" in sheet_name:
                            extractor = BaozhuangExtractor(sheet_name)
                        else:
                            # 默认提取器
                            extractor = BaozhuangExtractor(sheet_name)

                        if extractor:
                            try:
                                extractor.extract(ws, all_data)
                            except Exception as e:
                                st.error(f"处理工作表 '{sheet_name}' 时出错: {str(e)}")
                    
                    wb.close()
                    # 删除临时文件
                    os.unlink(tmp_path)

                except Exception as e:
                    st.error(f"❌ 处理文件 {uploaded_file.name} 时发生错误: {str(e)}")

            # 3. 输出结果 (替代原有的 output_file_path)
            status_text.text("处理完成，正在生成文件...")
            if all_data:
                output_buffer = save_to_output(all_data)
                
                st.success(f"✅ 处理完成！共提取有效记录 **{len(all_data)}** 条。")
                
                st.download_button(
                    label="📥 下载结果文件 (生产车间统计数据收集.xlsx)",
                    data=output_buffer,
                    file_name="生产车间统计数据收集.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.warning("⚠️ 未能提取到任何有效数据，请检查上传的文件格式是否正确。")

if __name__ == "__main__":
    main()
